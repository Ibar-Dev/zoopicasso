"""
Helper para operaciones con archivos (Excel, tickets, etc.)
"""
import os
from pathlib import Path
from typing import Optional, List, Dict
from openpyxl import load_workbook


class FileHelper:
    """Helper para operaciones con archivos de facturas"""
    
    @staticmethod
    def get_last_factura_file(facturas_dir: Path) -> Optional[Path]:
        """Obtiene el último archivo de factura generado"""
        if not facturas_dir.exists():
            return None
        
        # Buscar archivos .xlsx ordenados por fecha modificación
        xlsx_files = list(facturas_dir.glob("factura_*.xlsx"))
        if not xlsx_files:
            return None
        
        # Retorna el más reciente
        return max(xlsx_files, key=os.path.getmtime)
    
    @staticmethod
    def get_factura_numero_from_filename(filename: str) -> str:
        """Extrae número de factura del nombre del archivo
        Ejemplo: factura_2024_001.xlsx -> 2024-001"""
        # factura_2024_001.xlsx -> 2024-001
        base_name = Path(filename).stem  # factura_2024_001
        parts = base_name.split('_')  # ['factura', '2024', '001']
        if len(parts) >= 3:
            return f"{parts[1]}-{parts[2]}"
        return None
    
    @staticmethod
    def read_excel_total(excel_path: Path) -> Optional[float]:
        """Lee el total final del Excel"""
        try:
            wb = load_workbook(str(excel_path), data_only=True)
            ws = wb.active
            
            # Buscar la fila que contiene "TOTAL" en negrita
            # Generalmente está en la última fila con valor
            total = None
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "TOTAL" in str(cell.value).upper():
                        # La siguiente celda en la fila con datos numéricos es el total
                        for next_cell in row:
                            if next_cell.value and isinstance(next_cell.value, (int, float)):
                                total = float(next_cell.value)
                                break
            
            return total
        except Exception as e:
            print(f"Error leyendo Excel: {e}")
            return None
    
    @staticmethod
    def read_excel_lineas(excel_path: Path) -> List[Dict]:
        """Lee las líneas de factura del Excel"""
        try:
            wb = load_workbook(str(excel_path), data_only=True)
            ws = wb.active
            
            lineas = []
            in_lineas = False
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                # Buscar el inicio de la tabla de líneas (donde está "Concepto")
                if any(cell.value and isinstance(cell.value, str) and "Concepto" in str(cell.value) for cell in row):
                    in_lineas = True
                    continue
                
                if not in_lineas:
                    continue
                
                # Detener cuando encontramos "Subtotal" o "TOTAL"
                if any(cell.value and isinstance(cell.value, str) and ("Subtotal" in str(cell.value) or "TOTAL" in str(cell.value)) for cell in row):
                    break
                
                # Extraer datos de línea (Concepto, Cantidad, Precio, Total)
                values = [cell.value for cell in row]
                
                # Verificar que la fila tiene datos válidos (al menos 4 columnas)
                if len(values) >= 4 and values[0] and isinstance(values[2], (int, float)):
                    linea = {
                        'concepto': str(values[0]),
                        'categoria': str(values[1]) if values[1] else '',
                        'cantidad': int(values[2]) if isinstance(values[2], int) else float(values[2]),
                        'precio_unitario': float(values[3]) if isinstance(values[3], (int, float)) else 0.0,
                        'total': float(values[4]) if len(values) > 4 and isinstance(values[4], (int, float)) else 0.0,
                    }
                    lineas.append(linea)
            
            return lineas
        except Exception as e:
            print(f"Error leyendo líneas del Excel: {e}")
            return []
    
    @staticmethod
    def read_excel_cliente(excel_path: Path) -> Dict:
        """Lee datos del cliente del Excel"""
        try:
            wb = load_workbook(str(excel_path), data_only=True)
            ws = wb.active
            
            cliente_data = {'nombre': '', 'nif': ''}
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                for i, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        if "Nombre" in cell and i + 1 < len(row):
                            cliente_data['nombre'] = str(row[i + 1]) if row[i + 1] else ''
                        elif "NIF" in cell and i + 1 < len(row):
                            cliente_data['nif'] = str(row[i + 1]) if row[i + 1] else ''
            
            return cliente_data
        except Exception as e:
            print(f"Error leyendo datos del cliente del Excel: {e}")
            return {'nombre': '', 'nif': ''}
    
    @staticmethod
    def verify_factura_file_exists(facturas_dir: Path, numero_factura: str) -> bool:
        """Verifica si existe el archivo de factura"""
        # numero_factura formato: "2024-001"
        parts = numero_factura.split('-')
        if len(parts) != 2:
            return False
        
        expected_filename = f"factura_{parts[0]}_{parts[1]}.xlsx"
        expected_path = facturas_dir / expected_filename
        
        return expected_path.exists()
