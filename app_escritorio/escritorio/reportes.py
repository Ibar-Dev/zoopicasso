import logging
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.ventas_store import (
    historial_ventas,
    listar_ajustes_activos,
    resumen_ventas_activas,
    ventas_activas_detalle,
)

logger = logging.getLogger(__name__)

RUTA_REPORTES = Path(__file__).resolve().parent.parent / "reportes"

_AZUL_OSCURO = "1F4E79"
_AZUL_CLARO = "BDD7EE"
_GRIS_CLARO = "F2F2F2"
_BLANCO = "FFFFFF"

_THIN = Side(style="thin")
_BORDE = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold: bool = False, size: int = 11, color: str = "000000") -> Font:
    return Font(bold=bold, size=size, color=color)


def _cabecera(ws, fila: int, valores: list) -> None:
    for col_idx, valor in enumerate(valores, start=1):
        celda = ws.cell(row=fila, column=col_idx)
        celda.value = valor
        celda.font = _font(bold=True, size=10, color=_BLANCO)
        celda.fill = _fill(_AZUL_OSCURO)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        celda.border = _BORDE


def _fila(ws, fila: int, valores: list, alternar: bool = False, negrita: bool = False) -> None:
    color = _GRIS_CLARO if alternar else _BLANCO
    for col_idx, valor in enumerate(valores, start=1):
        celda = ws.cell(row=fila, column=col_idx)
        celda.value = valor
        celda.font = _font(bold=negrita)
        celda.border = _BORDE
        celda.fill = _fill(color)
        if isinstance(valor, float):
            celda.number_format = "#,##0.00"


def generar_reporte_mensual(anio_mes: str, destino: Path | None = None) -> Path:
    """Genera un Excel con resumen y detalle de las ventas activas del mes."""
    RUTA_REPORTES.mkdir(parents=True, exist_ok=True)
    destino = destino or (RUTA_REPORTES / f"reporte_ventas_{anio_mes}.xlsx")

    resumen = resumen_ventas_activas(anio_mes)
    detalle = ventas_activas_detalle(anio_mes)
    ajustes = listar_ajustes_activos(anio_mes)

    wb = openpyxl.Workbook()

    # ── Hoja resumen ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18

    ws["A1"] = f"RESUMEN DE VENTAS - {anio_mes}"
    ws["A1"].font = _font(bold=True, size=14, color=_BLANCO)
    ws["A1"].fill = _fill(_AZUL_OSCURO)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 26

    fila = 3
    ws.cell(row=fila, column=1, value="Total del mes (neto)").font = _font(bold=True)
    celda = ws.cell(row=fila, column=2, value=resumen["total"])
    celda.number_format = "#,##0.00"
    celda.font = _font(bold=True)
    fila += 1
    ws.cell(row=fila, column=1, value="Total bruto").font = _font()
    ws.cell(row=fila, column=2, value=resumen["total_bruto"]).number_format = "#,##0.00"
    fila += 1
    ws.cell(row=fila, column=1, value="Ajustes manuales").font = _font()
    ws.cell(row=fila, column=2, value=resumen["ajuste_total"]).number_format = "#,##0.00"
    fila += 1
    ws.cell(row=fila, column=1, value="Cantidad de ventas").font = _font()
    ws.cell(row=fila, column=2, value=resumen["cantidad_ventas"])
    fila += 1
    ws.cell(row=fila, column=1, value="Efectivo").font = _font()
    ws.cell(row=fila, column=2, value=resumen["total_efectivo"]).number_format = "#,##0.00"
    fila += 1
    ws.cell(row=fila, column=1, value="Tarjeta").font = _font()
    ws.cell(row=fila, column=2, value=resumen["total_tarjeta"]).number_format = "#,##0.00"
    fila += 2

    ws.cell(row=fila, column=1, value="TOTALES POR CATEGORIA").font = _font(bold=True)
    ws.merge_cells(f"A{fila}:B{fila}")
    ws[f"A{fila}"].fill = _fill(_AZUL_CLARO)
    ws[f"B{fila}"].fill = _fill(_AZUL_CLARO)
    fila += 1
    _cabecera(ws, fila, ["Categoria", "Total"])
    fila += 1
    for i, (categoria, total) in enumerate(sorted(resumen["por_categoria"].items())):
        _fila(ws, fila, [categoria, total], alternar=(i % 2 == 1))
        fila += 1

    if ajustes:
        fila += 1
        ws.cell(row=fila, column=1, value="AJUSTES MANUALES").font = _font(bold=True)
        ws.merge_cells(f"A{fila}:B{fila}")
        ws[f"A{fila}"].fill = _fill(_AZUL_CLARO)
        ws[f"B{fila}"].fill = _fill(_AZUL_CLARO)
        fila += 1
        _cabecera(ws, fila, ["Fecha", "Monto"])
        fila += 1
        for i, ajuste in enumerate(ajustes):
            fecha = ajuste["created_at"][:10]
            _fila(ws, fila, [fecha, ajuste["monto"]], alternar=(i % 2 == 1))
            fila += 1

    # ── Hoja detalle ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Detalle")
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 22
    _cabecera(ws2, 1, ["Numero", "Fecha", "Categoria", "Cliente", "Monto", "Usuario"])
    for i, fila_venta in enumerate(detalle):
        _fila(
            ws2,
            i + 2,
            [
                fila_venta["numero_factura"],
                fila_venta["fecha_venta"],
                fila_venta["categoria"],
                fila_venta["cliente_nombre"],
                fila_venta["monto"],
                fila_venta["usuario"],
            ],
            alternar=(i % 2 == 1),
        )

    wb.save(destino)
    logger.info("Reporte mensual %s generado en %s", anio_mes, destino)
    return destino


def generar_reporte_historial(
    filas: list[dict],
    destino: Path | None = None,
    nombre: str = "historial",
) -> Path:
    """Exporta un historial de ventas filtrado a Excel."""
    RUTA_REPORTES.mkdir(parents=True, exist_ok=True)
    hoy = date.today().isoformat()
    destino = destino or (RUTA_REPORTES / f"{nombre}_{hoy}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial"
    anchos = [16, 14, 22, 22, 14, 14, 12]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[chr(64 + i)].width = ancho
    _cabecera(ws, 1, ["Numero", "Fecha", "Cliente", "Categorias", "Monto", "Metodo pago", "Estado"])
    for i, fila_venta in enumerate(filas):
        _fila(
            ws,
            i + 2,
            [
                fila_venta["numero_factura"],
                fila_venta["fecha_venta"],
                fila_venta["cliente_nombre"],
                fila_venta["categorias"],
                fila_venta["monto_lineas"],
                fila_venta["metodo_pago"],
                fila_venta["estado"],
            ],
            alternar=(i % 2 == 1),
        )
    wb.save(destino)
    logger.info("Historial exportado a %s (%d filas)", destino, len(filas))
    return destino