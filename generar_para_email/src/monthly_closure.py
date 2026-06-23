import logging
import os
import platform
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from src.ventas_store import (
    cerrar_mes_atomico,
    puede_hacer_cierre,
    registrar_cierre_diario,
    resumen_ventas_activas,
    resumen_ventas_dia,
    resumen_ventas_mañana,
    resumen_ventas_tarde,
)

logger = logging.getLogger(__name__)

_BASE = Path(__file__).resolve().parent.parent

def _obtener_carpeta_descargas() -> Path:
    """
    Detecta y retorna la carpeta de Descargas del sistema operativo.
    
    Soporta:
    - Windows (español): C:\\Users\\{user}\\Descargas
    - Windows (inglés): C:\\Users\\{user}\\Downloads
    - Linux/Mac (español): $HOME/Descargas
    - Linux/Mac (inglés): $HOME/Downloads
    
    Si ninguna carpeta existe, crea la carpeta "Descargas" en el home.
    """
    home = Path.home()
    
    # Intentar rutas en orden de preferencia
    rutas_candidatas = [
        home / "Descargas",        # Español
        home / "Downloads",         # Inglés
    ]
    
    # Retornar la primera que exista
    for ruta in rutas_candidatas:
        if ruta.exists() and ruta.is_dir():
            logger.info("📁 Carpeta de descargas detectada: %s", ruta)
            return ruta
    
    # Si ninguna existe, crear "Descargas" en el home
    descargas = home / "Descargas"
    descargas.mkdir(parents=True, exist_ok=True)
    logger.info("📁 Carpeta de descargas creada: %s", descargas)
    return descargas


def _ruta_cierres_dir() -> Path:
    """Ruta general para guardar cierres (fallback por defecto)."""
    valor = os.getenv("CIERRES_DIR", "").strip()
    if not valor:
        # Por defecto, usar la carpeta de Descargas del sistema operativo
        return _obtener_carpeta_descargas()
    ruta = Path(valor).expanduser()
    if not ruta.is_absolute():
        ruta = _BASE / ruta
    return ruta.resolve()


def _ruta_cierre_manana_dir() -> Path:
    """Ruta para guardar cierres de mañana (automáticos). Fallback: CIERRES_DIR."""
    valor = os.getenv("CIERRE_MANANA_DIR", "").strip()
    if not valor:
        return _ruta_cierres_dir()
    ruta = Path(valor).expanduser()
    if not ruta.is_absolute():
        ruta = _BASE / ruta
    return ruta.resolve()


def _ruta_cierre_tarde_dir() -> Path:
    """Ruta para guardar cierres de tarde (automáticos). Fallback: CIERRES_DIR."""
    valor = os.getenv("CIERRE_TARDE_DIR", "").strip()
    if not valor:
        return _ruta_cierres_dir()
    ruta = Path(valor).expanduser()
    if not ruta.is_absolute():
        ruta = _BASE / ruta
    return ruta.resolve()


def _ruta_cierre_dia_completo_dir() -> Path:
    """Ruta para guardar cierres de día completo (automáticos). Fallback: CIERRES_DIR."""
    valor = os.getenv("CIERRE_DIA_COMPLETO_DIR", "").strip()
    if not valor:
        return _ruta_cierres_dir()
    ruta = Path(valor).expanduser()
    if not ruta.is_absolute():
        ruta = _BASE / ruta
    return ruta.resolve()


def _ruta_cierre_mes_dir() -> Path:
    """Ruta para guardar cierres del mes (automáticos). Fallback: CIERRES_DIR."""
    valor = os.getenv("CIERRE_MES_DIR", "").strip()
    if not valor:
        return _ruta_cierres_dir()
    ruta = Path(valor).expanduser()
    if not ruta.is_absolute():
        ruta = _BASE / ruta
    return ruta.resolve()


RUTA_CIERRES = _ruta_cierres_dir()
RUTA_CIERRE_MANANA = _ruta_cierre_manana_dir()
RUTA_CIERRE_TARDE = _ruta_cierre_tarde_dir()
RUTA_CIERRE_DIA_COMPLETO = _ruta_cierre_dia_completo_dir()
RUTA_CIERRE_MES = _ruta_cierre_mes_dir()


def _generar_excel_cierre(anio_mes: str, resumen: dict, ruta_destino: Path | None = None) -> Path:
    """Genera Excel de cierre mensual con opción de ruta personalizada."""
    if ruta_destino is None:
        ruta_destino = RUTA_CIERRES
    ruta_destino.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo = ruta_destino / f"Cierre_del_mes_{anio_mes}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cierre {anio_mes}"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    ws["A1"] = "Cierre mensual de ganancias"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    ws["A3"] = "Periodo"
    ws["B3"] = anio_mes
    ws["A4"] = "Total activo"
    ws["B4"] = float(resumen["total"])
    ws["B4"].number_format = "#,##0.00"
    ws["A5"] = "Cantidad de ventas"
    ws["B5"] = int(resumen["cantidad_ventas"])

    ws["A7"] = "Categoría"
    ws["B7"] = "Total"
    ws["A7"].font = Font(bold=True)
    ws["B7"].font = Font(bold=True)
    ws["A7"].fill = PatternFill("solid", fgColor="D9E1F2")
    ws["B7"].fill = PatternFill("solid", fgColor="D9E1F2")

    fila = 8
    for categoria, total in sorted(resumen["por_categoria"].items()):
        ws[f"A{fila}"] = categoria
        ws[f"B{fila}"] = float(total)
        ws[f"B{fila}"].number_format = "#,##0.00"
        fila += 1

    wb.save(archivo)
    if not archivo.exists() or archivo.stat().st_size == 0:
        raise OSError(f"No se pudo verificar el Excel de cierre: {archivo}")

    return archivo


def _generar_excel_cierre_dia(fecha: str, resumen: dict, tipo_cierre: str = "full_day", ruta_destino: Path | None = None) -> Path:
    """
    Genera Excel de cierre diario con nombre descriptivo según el tipo.
    
    tipo_cierre puede ser: "morning", "afternoon", "full_day"
    ruta_destino: ruta donde guardar el archivo (opcional, usa RUTA_CIERRES por defecto)
    """
    if ruta_destino is None:
        ruta_destino = RUTA_CIERRES
    ruta_destino.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # Nombres descriptivos según tipo de cierre
    nombres_por_tipo = {
        "morning": "Cierre_de_la_mañana",
        "afternoon": "Cierre_de_la_tarde",
        "full_day": "Cierre_del_dia_completo",
    }
    
    nombre_tipo = nombres_por_tipo.get(tipo_cierre, "Cierre_diario")
    archivo = ruta_destino / f"{nombre_tipo}_{fecha}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cierre {fecha}"

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    ws["A1"] = "Cierre diario de ganancias"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A1:B1")

    ws["A3"] = "Fecha"
    ws["B3"] = fecha
    ws["A4"] = "Total"
    ws["B4"] = float(resumen["total"])
    ws["B4"].number_format = "#,##0.00"
    ws["A5"] = "Cantidad de ventas"
    ws["B5"] = int(resumen["cantidad_ventas"])

    ws["A7"] = "Categoría"
    ws["B7"] = "Total"
    ws["A7"].font = Font(bold=True)
    ws["B7"].font = Font(bold=True)
    ws["A7"].fill = PatternFill("solid", fgColor="D9E1F2")
    ws["B7"].fill = PatternFill("solid", fgColor="D9E1F2")

    fila = 8
    for categoria, total in sorted(resumen["por_categoria"].items()):
        ws[f"A{fila}"] = categoria
        ws[f"B{fila}"] = float(total)
        ws[f"B{fila}"].number_format = "#,##0.00"
        fila += 1

    wb.save(archivo)
    if not archivo.exists() or archivo.stat().st_size == 0:
        raise OSError(f"No se pudo verificar el Excel de cierre diario: {archivo}")

    return archivo


def cerrar_mes(usuario: str) -> tuple[dict, Path | None]:
    """Cierra el mes activo: archiva ventas, genera Excel y registra el cierre.

    Devuelve (metadata, Path_al_excel) o (metadata, None) si no hay ventas.
    """
    anio_mes = datetime.now().strftime("%Y-%m")
    resumen = resumen_ventas_activas(anio_mes)

    if resumen["cantidad_ventas"] == 0:
        return {
            "ok": True,
            "anio_mes": anio_mes,
            "cantidad_ventas": 0,
            "total": 0.0,
            "mensaje": "No hay ventas activas para cerrar en este mes.",
        }, None

    archivo = _generar_excel_cierre(anio_mes, resumen, RUTA_CIERRE_MES)
    cierre_id = f"{anio_mes}-{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S%f')}"
    archived_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    actualizadas = cerrar_mes_atomico(
        anio_mes=anio_mes,
        cierre_id=cierre_id,
        archived_at=archived_at,
        usuario=(usuario or "").strip(),
        total=float(resumen["total"]),
        archivo_excel=str(archivo),
    )
    logger.info(
        "Cierre mensual. usuario=%s periodo=%s ventas=%d total=%.2f",
        usuario, anio_mes, actualizadas, float(resumen["total"]),
    )
    return {
        "ok": True,
        "anio_mes": anio_mes,
        "cantidad_ventas": int(actualizadas),
        "total": float(resumen["total"]),
        "mensaje": "Cierre mensual completado correctamente.",
    }, archivo


def _cerrar_dia_generico(usuario: str, tipo_cierre: str, resumen: dict, fecha: str, anio_mes: str, ruta_destino: Path | None = None) -> tuple[dict, Path | None]:
    """Función genérica para generar cierre y registrarlo. Valida secuencia.
    
    ruta_destino: ruta donde guardar el archivo (opcional, usa RUTA_CIERRES por defecto)
    """
    from src.ventas_store import puede_hacer_cierre
    
    puede, motivo = puede_hacer_cierre(tipo_cierre, fecha)
    if not puede:
        return {
            "ok": False,
            "fecha": fecha,
            "cantidad_ventas": 0,
            "total": 0.0,
            "mensaje": motivo,
        }, None

    if resumen["cantidad_ventas"] == 0:
        return {
            "ok": True,
            "fecha": fecha,
            "cantidad_ventas": 0,
            "total": 0.0,
            "mensaje": f"No hay ventas para cerrar en {resumen.get('periodo', tipo_cierre)}.",
        }, None

    archivo = _generar_excel_cierre_dia(fecha, resumen, tipo_cierre, ruta_destino)
    cierre_id = f"{tipo_cierre}-{fecha}-{datetime.now(timezone.utc).strftime('%H%M%S%f')}"
    created_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    registrar_cierre_diario(
        cierre_id=cierre_id,
        fecha=fecha,
        anio_mes=anio_mes,
        usuario=(usuario or "").strip(),
        created_at=created_at,
        total=float(resumen["total"]),
        cantidad_ventas=int(resumen["cantidad_ventas"]),
        archivo_excel=str(archivo),
        tipo_cierre=tipo_cierre,
    )
    logger.info(
        "Cierre %s. usuario=%s fecha=%s ventas=%d total=%.2f",
        tipo_cierre, usuario, fecha, resumen["cantidad_ventas"], float(resumen["total"]),
    )
    return {
        "ok": True,
        "fecha": fecha,
        "cantidad_ventas": int(resumen["cantidad_ventas"]),
        "total": float(resumen["total"]),
        "tipo_cierre": tipo_cierre,
        "mensaje": f"Cierre {tipo_cierre} completado correctamente.",
    }, archivo


def obtener_resumen_cierre_mañana(fecha: str) -> dict:
    """Obtiene el resumen de ventas de mañana sin registrar nada.
    
    Usado para FASE 3: mostrar ganancia bruta antes de preguntar por exportación.
    """
    resumen = resumen_ventas_mañana(fecha)
    return {
        "ok": True,
        "tipo_cierre": "morning",
        "periodo": "Mañana (06:00-14:00)",
        "dinero_bruto": float(resumen.get("total", 0)),
        "cantidad_ventas": int(resumen.get("cantidad_ventas", 0)),
        "por_categoria": resumen.get("por_categoria", {}),
    }


def obtener_resumen_cierre_tarde(fecha: str) -> dict:
    """Obtiene el resumen de ventas de tarde sin registrar nada.
    
    Usado para FASE 3: mostrar ganancia bruta antes de preguntar por exportación.
    """
    resumen = resumen_ventas_tarde(fecha)
    return {
        "ok": True,
        "tipo_cierre": "afternoon",
        "periodo": "Tarde (14:00-22:00)",
        "dinero_bruto": float(resumen.get("total", 0)),
        "cantidad_ventas": int(resumen.get("cantidad_ventas", 0)),
        "por_categoria": resumen.get("por_categoria", {}),
    }


def obtener_resumen_cierre_dia_completo(fecha: str) -> dict:
    """Obtiene el resumen de ventas del día completo sin registrar nada.
    
    Usado para FASE 3: mostrar ganancia bruta antes de preguntar por exportación.
    """
    resumen = resumen_ventas_dia(fecha)
    return {
        "ok": True,
        "tipo_cierre": "full_day",
        "periodo": "Día Completo (06:00-22:00)",
        "dinero_bruto": float(resumen.get("total", 0)),
        "cantidad_ventas": int(resumen.get("cantidad_ventas", 0)),
        "por_categoria": resumen.get("por_categoria", {}),
    }


def cerrar_dia(usuario: str) -> tuple[dict, Path | None]:
    """Genera informe Excel del día y registra el cierre. No archiva ventas. (Compatibilidad)
    
    Guarda en RUTA_CIERRE_DIA_COMPLETO si está configurada.
    """
    from src.ventas_store import resumen_ventas_dia
    fecha = datetime.now().strftime("%Y-%m-%d")
    anio_mes = datetime.now().strftime("%Y-%m")
    resumen = resumen_ventas_dia(fecha)
    return _cerrar_dia_generico(usuario, "full_day", resumen, fecha, anio_mes, RUTA_CIERRE_DIA_COMPLETO)


def cerrar_mañana(usuario: str) -> tuple[dict, Path | None]:
    """Cierre de mañana (06:00-14:00). Valida que no haya sido hecho hoy.
    
    Guarda en RUTA_CIERRE_MANANA si está configurada (automático).
    """
    from src.ventas_store import resumen_ventas_mañana
    fecha = datetime.now().strftime("%Y-%m-%d")
    anio_mes = datetime.now().strftime("%Y-%m")
    resumen = resumen_ventas_mañana(fecha)
    return _cerrar_dia_generico(usuario, "morning", resumen, fecha, anio_mes, RUTA_CIERRE_MANANA)


def cerrar_tarde(usuario: str) -> tuple[dict, Path | None]:
    """Cierre de tarde (14:00-22:00). Requiere que mañana esté hecho.
    
    Guarda en RUTA_CIERRE_TARDE si está configurada (automático).
    """
    from src.ventas_store import resumen_ventas_tarde
    fecha = datetime.now().strftime("%Y-%m-%d")
    anio_mes = datetime.now().strftime("%Y-%m")
    resumen = resumen_ventas_tarde(fecha)
    return _cerrar_dia_generico(usuario, "afternoon", resumen, fecha, anio_mes, RUTA_CIERRE_TARDE)


def cerrar_día_completo(usuario: str) -> tuple[dict, Path | None]:
    """Cierre del día completo. Requiere que mañana y tarde estén hechos.
    
    Guarda en RUTA_CIERRE_DIA_COMPLETO si está configurada (automático).
    """
    from src.ventas_store import resumen_ventas_dia
    fecha = datetime.now().strftime("%Y-%m-%d")
    anio_mes = datetime.now().strftime("%Y-%m")
    resumen = resumen_ventas_dia(fecha)
    return _cerrar_dia_generico(usuario, "full_day", resumen, fecha, anio_mes, RUTA_CIERRE_DIA_COMPLETO)
